import pandas as pd
from openpyxl import load_workbook

import tempfile

class Roster(object):
    """ Build this class.
    """

    # Probable data mapping:
    #    Roster: { [Student Name] -> Workbook:[row number] }
    #    Roster: { [Student Name] -> Workbook:[Worksheet name] }
    #    Roster: { [Student Name] -> { Roster:[Assignment number] -> [Grade] } }
    # 
    # Possible data mapping:
    #    Roster: { [Heading Name] -> Workbook:[Column Letter] }
    #    Roster: { [Student Name] -> [ID] }
    #    Roster: { [Student Name] -> [Class Grade formula] }
    


    def __init__(self, filename):
        self.__recalc_required__ = True
        self.__reset__(filename)
        self.file_extension = '.' + filename.split('.')[-1]

    def __save_tmp_reload__(self):
        tmp_xls = tempfile.NamedTemporaryFile(suffix=self.file_extension).name

        print("tmp_xls: " + str(tmp_xls))
        self.wb.save(tmp_xls)
        self.wb.close()

        self.__reset__(tmp_xls)

    def __reset__(self, filename):
        self.wb = load_workbook(filename)
        self.wb_data = load_workbook(filename, data_only=True)

        self.name_to_df_index = {}
        self.students_info = {}

        data = self.wb["Roster"].values
        cols = next(data)
        self.df_roster = pd.DataFrame(data, columns=cols)

        data = self.wb_data["Roster"].values
        cols = next(data)
        self.df_roster_data = pd.DataFrame(data, columns=cols)
        self.__recalc_required__ = False

        self.build_student_names()



    def build_student_names(self):
    # Get names of students from the excel workbook main worksheet ("Roster")
    # Row 1: Headers
    # Col A: ID, numbered sequentially starting at ID == 1
    # Col B: First Name
    # Col C: Last name
    # Col D: Class Grade average computed from each sheet, assuming that sheets are named ==
    # Student_[ID]

    # Returns: List of student names in the format "[First Name][space][Last Name]"

        self.student_names = []

        index = 0
        while (index < len(self.df_roster)) and (str(self.df_roster["First Name"][index]) != "None"):

            first_name = self.df_roster["First Name"][index]
            last_name = self.df_roster["Last Name"][index]

            print("build_student_names: first_name: " + str(first_name))
            print("build_student_names: last_name: " + str(last_name))

            full_name = str(first_name) + " " + str(last_name)

            self.student_names.append(full_name)

            self.name_to_df_index[full_name] = index

            index += 1

    def get_student_names(self):
        return self.student_names


    def get_student(self, student_name):
    # Return a dict where key -> value correspondences are
    # "id" -> [Integer ID]
    # "grades" -> [Grades as a <pandas.Series>]
    # student["grades"][[assignment number]] == [Integer grade on that assignment]
        student_info = {}
        df_index = self.name_to_df_index[student_name]
        student_info["id"] = int(self.df_roster["ID"][df_index])

        worksheet = "Student_" + str(student_info["id"])
        print("worksheet is " + worksheet)

        sheet = self.wb[worksheet]

        grades = []
        colx = 2
        rowx = 6
        while (sheet.cell(row=rowx, column=colx).value) != None:
            grade = sheet.cell(row=rowx, column=colx).value
            grades.append(grade)
            rowx += 1

        student_info["grades"] = pd.Series(grades)

        self.students_info[student_name] = student_info
        
        return student_info
    

    def class_average(self):
    # Return the average of each students grade average from the current, updated data
    # Original/pre-update data from the "Roster" sheet, "Class Grade" column

        if self.__recalc_required__:
            self.__save_tmp_reload__()

        print("Student names post reset: " + str(self.get_student_names()))
        
        grade_total = 0
        index = 0
        while index < len(self.df_roster_data):
            grade = self.df_roster_data["Class Grade"][index]
            index += 1

            print("Grade == " + str(grade))
            grade_total += grade

        print("grade_total: " + str(grade_total))

        self.class_average_grade = grade_total / index

        return self.class_average_grade


    def save(self, file_path):
    # Save the updated data to a workbook named [file_path]
        self.wb.save(file_path)


    def __renumber__(self, start_row, change_by):
        row_x = start_row
        id_col_x = 1
        formula_col_x = 4
        fname_col_x = 2

        sheet = self.wb["Roster"]

        while (sheet.cell(row=row_x, column=id_col_x).value != None):
            old_id = sheet.cell(row=row_x, column=id_col_x).value
            new_id = old_id + change_by
            sheet.cell(row=row_x, column=id_col_x).value = str(new_id)
            
            first_name = sheet.cell(row=row_x, column=fname_col_x).value

            old_formula = sheet.cell(row=row_x, column=formula_col_x).value
            old_sheet = "Student_" + str(old_id)
            new_sheet = "Student_" + str(new_id)

            new_formula = old_formula.replace(old_sheet, new_sheet)

            print("first name: " + str(first_name))
            print("old_sheet: " + str(old_sheet))
            print("new_sheet: " + str(new_sheet))

            print("new id: " + str(sheet.cell(row=row_x, column=id_col_x).value))

            print("old_formula: " + str(old_formula))
            print("new_formula: " + str(new_formula))

            sheet.cell(row=row_x, column=formula_col_x).value = new_formula

            indiv_sheet = self.wb[old_sheet]
            id_col = 2
            id_row = 1
            id_cell = indiv_sheet.cell(row=id_row, column=id_col)
            id_cell.value = str(new_id)

            indiv_sheet.title = new_sheet

            row_x += 1




    def delete_student(self, student_name, renumber=True):
    # Delete student_name [[First Name][space][Last Name]] from the updated / class-state workbook data
    # Initial pass at sequence of steps:
    #    (1) Delete row
    #    (2) Delete the worksheet corresponding to the deleted student
    #
    #    (3) If "renumber" argument is true
    #            Decrement, by 1, all higher ID numbers than that deleted
    #            Update (grade) formulas so source worksheet name uses new ID suffix if changed
    #            Decrement the ID part of worksheet names ( Student_[ID] ).


        roster_sheet = self.wb["Roster"]
        df_index = self.name_to_df_index[student_name]
        sheet_row = df_index + 2
        roster_sheet.delete_rows(idx=sheet_row, amount=1)

        old_end_row = len(self.df_roster) + 1
        print("Deleting old end row: " + str(old_end_row))
        roster_sheet.delete_rows(idx=old_end_row, amount=1)

        student_info = self.get_student(student_name)
        id = student_info["id"]
        worksheet_name = "Student_" + str(id)
        worksheet = self.wb[worksheet_name]
        self.wb.remove(worksheet)

        # data = self.wb["Roster"].values
        # cols = next(data)
        # self.df_roster = pd.DataFrame(data, columns=cols)

        # self.wb_data = None
        # self.df_roster_data = None
        # self.__recalc_required__ = True

        self.student_names.remove(student_name)

        del self.name_to_df_index[student_name]
        del self.students_info[student_name]

        if renumber:
            self.__renumber__(start_row=sheet_row, change_by=-1)

        data = self.wb["Roster"].values
        cols = next(data)
        self.df_roster = pd.DataFrame(data, columns=cols)

        self.wb_data = None
        self.df_roster_data = None
        self.__recalc_required__ = True

        self.build_student_names()




    def add_student(self, student):
    # Optional?
    # Analagous to delete_student(), exceptions include:
    #     New students worksheet should be added before "Roster" sheet is updated
    #     Assumption: After current students
        pass
